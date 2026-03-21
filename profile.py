"""
tools/profile.py — Enhanced File Profiler for ARIA-LAB v2

Supports: CSV, JSON, TXT, MD, PDF, XLSX, YAML, Python files
For each: summary stats, schema inference, content preview, anomaly detection.

Usage:
    python tools/profile.py data.csv
    python tools/profile.py inputs/
    python tools/profile.py data.json --full
"""

from __future__ import annotations

import argparse
import ast
import csv
import io
import json
import logging
import math
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".csv", ".json", ".txt", ".md", ".pdf", ".xlsx", ".yaml", ".yml", ".py"}


# ──────────────────────────────────────────────
# Data Models
# ──────────────────────────────────────────────

@dataclass
class ColumnProfile:
    name:         str
    dtype:        str         # string | integer | float | boolean | mixed | null
    count:        int
    null_count:   int
    unique_count: int
    sample_values: list[str]  = field(default_factory=list)
    min_val:      Optional[Any] = None
    max_val:      Optional[Any] = None
    mean_val:     Optional[float] = None
    anomalies:    list[str]   = field(default_factory=list)


@dataclass
class FileProfile:
    path:           str
    extension:      str
    size_bytes:     int
    encoding:       str          = "utf-8"
    line_count:     int          = 0
    word_count:     int          = 0
    char_count:     int          = 0
    schema:         dict         = field(default_factory=dict)
    columns:        list[ColumnProfile] = field(default_factory=list)
    record_count:   int          = 0
    preview:        str          = ""
    summary:        str          = ""
    anomalies:      list[str]    = field(default_factory=list)
    metadata:       dict         = field(default_factory=dict)
    error:          str          = ""

    def to_dict(self) -> dict:
        d = asdict(self)
        d["columns"] = [asdict(c) for c in self.columns]
        return d


# ──────────────────────────────────────────────
# Type Inference
# ──────────────────────────────────────────────

def _infer_type(values: list[str]) -> str:
    non_null = [v for v in values if v and v.lower() not in ("", "null", "none", "na", "n/a")]
    if not non_null:
        return "null"
    ints   = sum(1 for v in non_null if re.fullmatch(r"-?\d+", v))
    floats = sum(1 for v in non_null if re.fullmatch(r"-?\d+\.\d+|-?\d+[eE][+-]?\d+", v))
    bools  = sum(1 for v in non_null if v.lower() in ("true","false","yes","no","1","0"))
    n = len(non_null)
    if ints / n > 0.9:
        return "integer"
    if (ints + floats) / n > 0.9:
        return "float"
    if bools / n > 0.9:
        return "boolean"
    if all(len(v) == len(non_null[0]) for v in non_null[:20]) and not any(c.isalpha() for v in non_null[:5] for c in v):
        return "code/id"
    return "string"


def _numeric_stats(values: list[str]) -> tuple[Optional[float], Optional[float], Optional[float]]:
    nums: list[float] = []
    for v in values:
        try:
            nums.append(float(v.replace(",", "")))
        except (ValueError, AttributeError):
            pass
    if not nums:
        return None, None, None
    return min(nums), max(nums), sum(nums) / len(nums)


def _detect_anomalies_column(name: str, values: list[str], dtype: str) -> list[str]:
    anomalies: list[str] = []
    null_rate = sum(1 for v in values if not v or v.lower() in ("null","none","na")) / max(len(values), 1)
    if null_rate > 0.3:
        anomalies.append(f"High null rate: {null_rate:.0%}")
    if dtype in ("integer", "float"):
        nums = [float(v.replace(",","")) for v in values if re.fullmatch(r"[-\d.,eE+]+", v)]
        if nums:
            q1, q3 = sorted(nums)[len(nums)//4], sorted(nums)[3*len(nums)//4]
            iqr = q3 - q1
            outliers = [n for n in nums if n < q1 - 3*iqr or n > q3 + 3*iqr]
            if outliers:
                anomalies.append(f"{len(outliers)} outlier(s) detected (IQR method)")
    return anomalies


# ──────────────────────────────────────────────
# Format-Specific Profilers
# ──────────────────────────────────────────────

def _profile_csv(path: Path) -> FileProfile:
    profile = FileProfile(path=str(path), extension=".csv", size_bytes=path.stat().st_size)
    try:
        # Detect encoding
        raw = path.read_bytes()
        profile.char_count = len(raw)
        encoding = "utf-8"
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                raw.decode(enc)
                encoding = enc
                break
            except UnicodeDecodeError:
                continue
        profile.encoding = encoding

        text = raw.decode(encoding, errors="replace")

        # Detect dialect
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader   = csv.DictReader(io.StringIO(text), dialect=dialect)
        rows     = list(reader)
        headers  = reader.fieldnames or []

        profile.record_count = len(rows)
        profile.line_count   = len(rows) + 1
        profile.schema       = {h: "unknown" for h in headers}

        SAMPLE_N = 200
        for col in headers:
            values = [str(row.get(col, "")) for row in rows[:SAMPLE_N]]
            dtype  = _infer_type(values)
            mn, mx, mean = _numeric_stats(values) if dtype in ("integer","float") else (None, None, None)
            profile.schema[col] = dtype
            profile.columns.append(ColumnProfile(
                name=col,
                dtype=dtype,
                count=len(values),
                null_count=sum(1 for v in values if not v),
                unique_count=len(set(values)),
                sample_values=[v for v in values[:5] if v],
                min_val=mn, max_val=mx, mean_val=mean,
                anomalies=_detect_anomalies_column(col, values, dtype),
            ))

        # Global anomalies
        if profile.record_count == 0:
            profile.anomalies.append("File is empty (no data rows)")
        if len(headers) > 50:
            profile.anomalies.append(f"High column count: {len(headers)}")

        profile.preview = text[:500]
        profile.summary = (
            f"CSV with {profile.record_count:,} rows × {len(headers)} columns. "
            f"Types: {dict(list(profile.schema.items())[:5])}…"
        )
    except Exception as e:
        profile.error = str(e)
        logger.error("CSV profile failed: %s", e)
    return profile


def _profile_json(path: Path) -> FileProfile:
    profile = FileProfile(path=str(path), extension=".json", size_bytes=path.stat().st_size)
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        data = json.loads(text)
        profile.char_count = len(text)
        profile.line_count = text.count("\n")

        def _schema_of(obj: Any, depth: int = 0) -> Any:
            if depth > 3:
                return type(obj).__name__
            if isinstance(obj, dict):
                return {k: _schema_of(v, depth+1) for k, v in list(obj.items())[:20]}
            elif isinstance(obj, list):
                if not obj:
                    return []
                return [_schema_of(obj[0], depth+1)]
            return type(obj).__name__

        if isinstance(data, list):
            profile.record_count = len(data)
            if data and isinstance(data[0], dict):
                profile.schema = _schema_of(data[0])
                # Column profiles from first 200 records
                for key in data[0].keys():
                    vals = [str(r.get(key, "")) for r in data[:200] if isinstance(r, dict)]
                    dtype = _infer_type(vals)
                    profile.schema[key] = dtype
        else:
            profile.schema = _schema_of(data)

        preview_data = json.dumps(data, indent=2)[:500]
        profile.preview = preview_data

        # Anomalies
        if isinstance(data, list) and len(data) == 0:
            profile.anomalies.append("Empty array")

        profile.summary = (
            f"JSON {'array' if isinstance(data, list) else 'object'}"
            + (f" with {len(data):,} records" if isinstance(data, list) else "")
            + f". Schema keys: {list(profile.schema.keys())[:8]}"
        )
    except Exception as e:
        profile.error = str(e)
    return profile


def _profile_text(path: Path) -> FileProfile:
    profile = FileProfile(path=str(path), extension=path.suffix, size_bytes=path.stat().st_size)
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        profile.line_count = text.count("\n")
        profile.char_count = len(text)
        profile.word_count = len(text.split())
        profile.preview    = text[:500]

        # Anomalies for text
        if profile.word_count < 10:
            profile.anomalies.append("Very short document (<10 words)")
        if profile.char_count > 1_000_000:
            profile.anomalies.append("Very large file (>1MB text)")

        profile.summary = (
            f"Text file: {profile.line_count:,} lines, "
            f"{profile.word_count:,} words, {profile.char_count:,} chars."
        )
    except Exception as e:
        profile.error = str(e)
    return profile


def _profile_python(path: Path) -> FileProfile:
    profile = _profile_text(path)
    profile.extension = ".py"
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        tree = ast.parse(text)

        classes   = [n.name for n in ast.walk(tree) if isinstance(n, ast.ClassDef)]
        functions = [n.name for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)]
        imports_  = [ast.dump(n) for n in ast.walk(tree)
                     if isinstance(n, (ast.Import, ast.ImportFrom))]

        profile.metadata = {
            "classes":   classes[:20],
            "functions": functions[:30],
            "class_count": len(classes),
            "function_count": len(functions),
            "import_count":   len(imports_),
        }
        profile.schema = {
            "classes":   "list[str]",
            "functions": "list[str]",
        }
        profile.summary = (
            f"Python module: {len(classes)} class(es), {len(functions)} function(s), "
            f"{len(imports_)} import(s). {profile.line_count:,} lines."
        )

        # Syntax check
    except SyntaxError as e:
        profile.anomalies.append(f"Syntax error: {e}")
    except Exception as e:
        profile.error = str(e)
    return profile


def _profile_xlsx(path: Path) -> FileProfile:
    profile = FileProfile(path=str(path), extension=".xlsx", size_bytes=path.stat().st_size)
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        profile.metadata = {"sheets": sheet_names}

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True, max_row=201))
        if not rows:
            profile.summary = "Empty XLSX"
            return profile

        headers = [str(c) if c is not None else "" for c in rows[0]]
        data    = rows[1:]
        profile.record_count = len(data)
        profile.schema = {h: "unknown" for h in headers}

        for i, col in enumerate(headers):
            vals = [str(row[i]) if i < len(row) and row[i] is not None else "" for row in data[:200]]
            dtype = _infer_type(vals)
            profile.schema[col] = dtype

        profile.summary = (
            f"XLSX with {len(sheet_names)} sheet(s). "
            f"Active sheet: {profile.record_count:,} rows × {len(headers)} columns."
        )
    except ImportError:
        profile.error = "openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        profile.error = str(e)
    return profile


def _profile_yaml(path: Path) -> FileProfile:
    profile = FileProfile(path=str(path), extension=path.suffix, size_bytes=path.stat().st_size)
    try:
        import yaml  # type: ignore
        text = path.read_text(encoding="utf-8", errors="replace")
        data = yaml.safe_load(text)
        profile.char_count = len(text)
        profile.line_count = text.count("\n")
        profile.preview    = text[:500]

        def _walk_schema(obj: Any, depth: int = 0) -> Any:
            if depth > 3:
                return type(obj).__name__
            if isinstance(obj, dict):
                return {k: _walk_schema(v, depth+1) for k in list(obj.keys())[:20]}
            elif isinstance(obj, list):
                return [_walk_schema(obj[0], depth+1)] if obj else []
            return type(obj).__name__

        profile.schema = _walk_schema(data) if data else {}
        profile.summary = f"YAML file: {profile.line_count:,} lines."
    except ImportError:
        profile.error = "pyyaml not installed. Run: pip install pyyaml"
    except Exception as e:
        profile.error = str(e)
    return profile


def _profile_pdf(path: Path) -> FileProfile:
    profile = FileProfile(path=str(path), extension=".pdf", size_bytes=path.stat().st_size)
    try:
        # Try pdfplumber first, then pypdf
        text = ""
        try:
            import pdfplumber  # type: ignore
            with pdfplumber.open(path) as pdf:
                profile.metadata["page_count"] = len(pdf.pages)
                text = "\n".join(p.extract_text() or "" for p in pdf.pages[:10])
        except ImportError:
            try:
                from pypdf import PdfReader  # type: ignore
                reader = PdfReader(path)
                profile.metadata["page_count"] = len(reader.pages)
                text = "\n".join(p.extract_text() or "" for p in reader.pages[:10])
            except ImportError:
                profile.error = "No PDF library. pip install pdfplumber"
                return profile

        profile.char_count  = len(text)
        profile.word_count  = len(text.split())
        profile.line_count  = text.count("\n")
        profile.preview     = text[:500]
        pages = profile.metadata.get("page_count", "?")
        profile.summary = (
            f"PDF with {pages} page(s). "
            f"Extracted {profile.word_count:,} words from first 10 pages."
        )
    except Exception as e:
        profile.error = str(e)
    return profile


# ──────────────────────────────────────────────
# Main Profiler
# ──────────────────────────────────────────────

class FileProfiler:
    """Profiles any supported file type."""

    PROFILERS = {
        ".csv":  _profile_csv,
        ".json": _profile_json,
        ".txt":  _profile_text,
        ".md":   _profile_text,
        ".py":   _profile_python,
        ".xlsx": _profile_xlsx,
        ".xls":  _profile_xlsx,
        ".yaml": _profile_yaml,
        ".yml":  _profile_yaml,
        ".pdf":  _profile_pdf,
    }

    def profile(self, path: Path) -> FileProfile:
        ext = path.suffix.lower()
        fn  = self.PROFILERS.get(ext)
        if fn is None:
            p = FileProfile(path=str(path), extension=ext, size_bytes=path.stat().st_size)
            p.error = f"Unsupported file type: {ext}"
            return p
        try:
            return fn(path)
        except Exception as e:
            p = FileProfile(path=str(path), extension=ext, size_bytes=path.stat().st_size)
            p.error = str(e)
            return p

    def profile_directory(self, directory: Path) -> list[FileProfile]:
        profiles: list[FileProfile] = []
        for path in sorted(directory.rglob("*")):
            if path.is_file() and path.suffix.lower() in self.PROFILERS:
                logger.info("Profiling: %s", path.name)
                profiles.append(self.profile(path))
        return profiles


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────

def _cli() -> None:
    parser = argparse.ArgumentParser(
        description="ARIA-LAB File Profiler",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  python tools/profile.py data.csv
  python tools/profile.py inputs/
  python tools/profile.py data.json --full
  python tools/profile.py data.csv --out profile.json
""",
    )
    parser.add_argument("path",    type=Path, help="File or directory to profile")
    parser.add_argument("--full",  action="store_true", help="Show full profile including columns")
    parser.add_argument("--out",   type=Path,           help="Save profile to JSON file")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.WARNING,
        format="%(levelname)s: %(message)s",
    )

    profiler = FileProfiler()

    if not args.path.exists():
        print(f"Path not found: {args.path}", file=sys.stderr)
        sys.exit(1)

    if args.path.is_dir():
        profiles = profiler.profile_directory(args.path)
    else:
        profiles = [profiler.profile(args.path)]

    all_data = [p.to_dict() for p in profiles]

    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(json.dumps(all_data, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Profile saved to {args.out}")
        return

    for p in profiles:
        print(f"\n{'='*60}")
        print(f"  {p.path}")
        print(f"{'='*60}")
        if p.error:
            print(f"  ⚠ ERROR: {p.error}")
            continue
        size_str = f"{p.size_bytes:,} bytes"
        if p.size_bytes > 1024*1024:
            size_str = f"{p.size_bytes/1024/1024:.1f} MB"
        elif p.size_bytes > 1024:
            size_str = f"{p.size_bytes/1024:.1f} KB"
        print(f"  Size     : {size_str}")
        print(f"  Summary  : {p.summary}")
        if p.record_count:
            print(f"  Records  : {p.record_count:,}")
        if p.schema:
            print(f"  Schema   : {json.dumps(p.schema, ensure_ascii=False)[:120]}")
        if p.anomalies:
            print(f"  Anomalies:")
            for a in p.anomalies:
                print(f"    ⚠ {a}")
        if args.full and p.columns:
            print(f"\n  Columns ({len(p.columns)}):")
            for c in p.columns:
                anomaly_str = f" [!{'; '.join(c.anomalies)}]" if c.anomalies else ""
                print(f"    {c.name:<25} {c.dtype:<12} "
                      f"nulls={c.null_count} uniq={c.unique_count}{anomaly_str}")
        if p.preview:
            print(f"\n  Preview:")
            for line in p.preview.splitlines()[:10]:
                print(f"    {line[:100]}")


if __name__ == "__main__":
    _cli()
